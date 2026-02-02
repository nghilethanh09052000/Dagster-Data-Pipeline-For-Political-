from __future__ import annotations

import datetime
from collections.abc import Generator
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
import xlrd

if TYPE_CHECKING:
    from _typeshed import FileDescriptorOrPath


def save_parse_excel(
    file_path: Path, worksheet_idx: int = 0
) -> Generator[list[str], None, None]:
    """
    Safely decode and parse .xls or .xlsx to an iterable of string arrays.

    Parameters:
    file_path: the path of the ".xls" or ".xlsx" file
    worksheet_idx: the index (0 based) of the worksheet to read through, defaults to 0
    """

    if file_path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(filename=str(file_path), on_demand=True)
        sheet = workbook.sheet_by_index(worksheet_idx)

        for row_idx in range(sheet.nrows):
            row = []
            for cell in sheet.row(row_idx):
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)  # type: ignore
                        dt = datetime.datetime(
                            year=date_tuple[0],
                            month=date_tuple[1],
                            day=date_tuple[2],
                            hour=date_tuple[3],
                            minute=date_tuple[4],
                            second=date_tuple[5],
                        )
                        row.append(dt.isoformat(sep="T"))
                    except Exception:
                        row.append(str(cell.value))
                else:
                    row.append(str(cell.value))
            yield row

    elif file_path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(
            filename=str(file_path), read_only=True, data_only=True
        )
        sheetnames = workbook.sheetnames
        if worksheet_idx >= len(sheetnames):
            raise IndexError(
                f"Worksheet index {worksheet_idx} out of range for this file."
            )
        sheet = workbook[sheetnames[worksheet_idx]]

        for row in sheet.iter_rows(values_only=True):
            formatted_row = []
            for cell in row:
                if isinstance(cell, datetime.datetime | datetime.date):
                    formatted_row.append(cell.isoformat())
                else:
                    formatted_row.append(str(cell) if cell is not None else "")
            yield formatted_row

    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}")


def safe_readline_csv_like_file(
    file_path: FileDescriptorOrPath, encoding: str = "us-ascii"
) -> Generator[str]:
    """
    Safely decode and clean csv-like (character separated, new line indicates new row)
    from pesky NUL (0x00) bytes.

    The intended purpose of this function is to create a Generator that will be passed
    to `csv.Reader`.

    Parameters:
    file_path: the path of the csv-like file
    encoding: the encoding that should be used to open the file, defaults to "us-ascii"
              for legacy compability
    """

    # NOTE: there's a problem with 0xbf in one of the file, for the sake
    # of keeping it running on any encoding, we'll just use UTF-8
    # (most common) and replace the error characters with ""
    # (through `errors="ignore"`)
    with open(
        file=file_path, encoding=encoding, errors="ignore"
    ) as federal_bulk_download_file:
        for raw_row in federal_bulk_download_file:
            # NOTE: replace all NUL (0x00) ASCII somehow left out on
            # some of the FEC data files, making csv read error. Need
            # to replace it to "" manually
            cleaned_row = raw_row.replace("\x00", "")

            # TODO: add more characters to be cleaned if need be

            yield cleaned_row
