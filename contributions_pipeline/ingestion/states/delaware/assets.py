import csv
import glob
import os
import re
from collections.abc import Callable
from datetime import datetime
from pathlib import Path

import dagster as dg
import requests
from bs4 import BeautifulSoup
from bs4.element import Tag
from playwright.sync_api import sync_playwright
from psycopg import sql
from psycopg_pool import ConnectionPool

from contributions_pipeline.lib.fetch import stream_download_file_to_path
from contributions_pipeline.lib.file import (
    safe_readline_csv_like_file,
)
from contributions_pipeline.lib.ingest import (
    get_sql_count_query,
    get_sql_truncate_query,
    insert_parsed_file_to_landing_table,
)
from contributions_pipeline.resources import PostgresResource

# ===== Constants =====
# Base URL and Paths
DE_BASE_URL = "https://cfrs.elections.delaware.gov"
DE_CANDIDATES_BASE_URL = "https://elections.delaware.gov"
DE_DATA_PATH_PREFIX = "./states/delaware"

# Endpoints
DE_VIEW_REPORTS_URL = f"{DE_BASE_URL}/Public/ViewFiledReports"
DE_EXPORT_CSV_URL = f"{DE_BASE_URL}/Public/ExportFiledReportsToCsv"
DE_VIEW_RECEIPTS_URL = f"{DE_BASE_URL}/Public/ViewReceipts"
DE_EXPORT_CONTRIBUTIONS_URL = f"{DE_BASE_URL}/Public/ExportCSVNew"
DE_SEARCH_URL = f"{DE_BASE_URL}/Public/Search"
DE_EXPORT_COMMITTEES_URL = f"{DE_BASE_URL}/Public/ExporttoCsv"
DE_OTHER_SEARCH_URL = f"{DE_BASE_URL}/Public/OtherSearch"
DE_EXPORT_EXPENSES_URL = f"{DE_BASE_URL}/Public/ExportExpensestoCsv"
DE_CANDIDATES_LIST_URL = f"{DE_CANDIDATES_BASE_URL}/candidates/candidatelist/"

# Data Types
COMMITTEE_TYPES = ["01", "02", "03", "04", "05"]
OFFICE_TYPES = ["SO", "CO", "SB", "MO"]
DE_FIRST_YEAR_DATA_AVAILABLE = datetime(1980, 1, 1)


# Regex patterns for candidate pages
def get_candidate_page_pattern(year: int) -> str:
    """Generate year-specific regex pattern for candidate pages."""
    return rf'href="([^"]*?/([^/]+)_{year}\.shtml)"'


def get_xlsx_file_pattern(year: int) -> str:
    """Generate year-specific regex pattern for xlsx files."""
    return rf'href="([^"]*?([^/]+)_{year}\.xlsx)"'


# ===== Headers =====
# Filed Reports Headers
DE_FILED_REPORTS_HEADERS = [
    "filing_period",
    "filing_method",
    "cf_id",
    "committee_name",
    "committee_type",
    "original_report_filed_date",
    "latest_report_filed_date",
    "reporting_year",
    "office",
]

# Committees Headers
DE_COMMITTEES_HEADERS = [
    "committee_type",
    "cf_id",
    "committee_name",
    "office",
    "committee_status",
    "registered_date",
    "amendment_date",
    "tressurer_name",
    "tressurer_address",
    "empty_column",
]

# Contributions Headers
DE_CONTRIBUTIONS_HEADERS = [
    "contribution_date",
    "contributor_name",
    "contributor_address_line1",
    "contributor_address_line2",
    "contributor_city",
    "contributor_state",
    "contributor_zip",
    "contributor_type",
    "employer_name",
    "employer_occupation",
    "contribution_type",
    "contribution_amount",
    "cf_id",
    "receiving_committee",
    "filing_period",
    "office",
    "fixed_asset",
]

# Expenditures Headers
DE_EXPENDITURES_HEADERS = [
    "expenditure_date",
    "payee_name",
    "payee_address_line_1",
    "payee_address_line_2",
    "payee_city",
    "payee_state",
    "payee_zip",
    "payee_type",
    "amount_dollars",
    "cf_id",
    "committee_name",
    "expense_category",
    "expense_purpose",
    "expense_method",
    "filing_period",
    "fixed_asset",
    "empty_column",
]

# Candidates Headers
DE_CANDIDATES_HEADERS = [
    "name",
    "address",
    "party",
    "date_filed",
    "phone",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/135.0.0.0 Safari/537.36 Edg/135.0.0.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
    "image/avif,image/webp,image/apng,*/*;q=0.8,"
    "application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# Enhanced headers specifically for candidates (more modern browser-like)
CANDIDATES_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8,"
        "application/signed-exchange;v=b3;q=0.7"
    ),
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Accept-Language": "en-US,en;q=0.9",
    "Priority": "u=0, i",
    "Sec-Ch-Ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Microsoft Edge";v="138"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"macOS"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
}

# ===== Partition Definition =====
de_contributions_yearly_partition = dg.TimeWindowPartitionsDefinition(
    # Partition for year
    cron_schedule="0 0 1 1 *",
    # Make each of the partition to be yearly
    fmt="%Y",
    start=DE_FIRST_YEAR_DATA_AVAILABLE,
    end_offset=1,
)

de_candidates_yearly_partition = dg.TimeWindowPartitionsDefinition(
    # Partition for year
    cron_schedule="0 0 1 1 *",
    # Make each of the partition to be yearly
    fmt="%Y",
    start=DE_FIRST_YEAR_DATA_AVAILABLE,
    end_offset=1,
)


# ===== Common Functions =====
def de_insert_raw_file_to_landing_table(
    postgres_pool: ConnectionPool,
    table_name: str,
    table_columns_name: list[str],
    data_validation_callback: Callable[[list[str]], bool],
    category: str,
    data_files: list[str] | None = None,
    truncate_query: sql.Composed | None = None,
) -> dg.MaterializeResult:
    """
    Insert raw CSV data from previously downloaded files into the specified
    PostgreSQL landing table.

    Args:
        postgres_pool: Postgres connection pool
        table_name: Name of the landing table
        table_columns_name: List of column names
        data_validation_callback: Function to validate data rows
        category: The category of data (committees, contributions, etc.)
        data_files: Optional list of specific file paths to process
        truncate_query: Optional custom truncate query

    Returns:
        MaterializeResult with metadata about the operation
    """
    logger = dg.get_dagster_logger(name=f"{category}_insert")

    if data_files is None:
        # Use automatic file discovery for non-contributions
        base_path = Path(DE_DATA_PATH_PREFIX)
        data_files = []

        glob_path = base_path / category
        if glob_path.exists():
            # Look for CSV files only
            csv_files = glob.glob(str(glob_path / f"*_{category}.csv"))
            data_files.extend(csv_files)
    else:
        # Use the provided data_files list
        data_files = [str(f) for f in data_files]

    logger.info(f"Category: {category} | Found {len(data_files)} files.")

    if truncate_query is None:
        truncate_query = get_sql_truncate_query(table_name=table_name)

    with (
        postgres_pool.connection() as pg_connection,
        pg_connection.cursor() as pg_cursor,
    ):
        logger.info(f"Executing truncate for table {table_name}.")
        pg_cursor.execute(query=truncate_query)

        try:
            for data_file in data_files:
                if not os.path.exists(data_file):
                    logger.warning(f"File {data_file} does not exist, skipping...")
                    continue

                logger.info(f"Inserting from: {data_file}")

                # Process CSV files
                logger.info(f"Processing CSV file: {data_file}")

                data_type_lines_generator = safe_readline_csv_like_file(
                    data_file, encoding="utf-8"
                )

                parsed_data_type_file = csv.reader(
                    data_type_lines_generator, delimiter=","
                )

                # Skip header
                next(parsed_data_type_file)

                insert_parsed_file_to_landing_table(
                    pg_cursor=pg_cursor,
                    csv_reader=parsed_data_type_file,
                    table_name=table_name,
                    table_columns_name=table_columns_name,
                    row_validation_callback=data_validation_callback,
                )

        except Exception as e:
            logger.error(f"Error while processing {category}: {e}")
            raise e

        count_query = get_sql_count_query(table_name=table_name)
        count_cursor_result = pg_cursor.execute(query=count_query).fetchone()
        row_count = int(count_cursor_result[0]) if count_cursor_result else 0

        logger.info(f"Inserted {row_count} rows into {table_name}")

        return dg.MaterializeResult(
            metadata={"dagster/table_name": table_name, "dagster/row_count": row_count}
        )


# ===== Helper Functions =====
def validate_and_fix_date(date_value: str, year: int) -> str:
    """
    Validate a date string and fix invalid dates by
    replacing them with the first day of the year.

    Args:
        date_value: Date string in MM/DD/YYYY format
        year: Year to use for fixing invalid dates

    Returns:
        Valid date string in MM/DD/YYYY format
    """
    # Validate year parameter
    if not isinstance(year, int) or year < 1900 or year > 2100:
        print(f"Invalid year parameter: {year}, using 2020 as fallback")
        year = 2020

    if not date_value or not str(date_value).strip():
        print(f"Empty date value, replacing with 01/01/{year}")
        return f"01/01/{year}"

    date_str = str(date_value).strip()

    # Check if it matches MM/DD/YYYY pattern
    date_pattern = r"^([0-9]{1,2})/([0-9]{1,2})/([0-9]{4})$"
    match = re.match(date_pattern, date_str)

    if not match:
        print(f"Invalid date format: '{date_str}', replacing with 01/01/{year}")
        return f"01/01/{year}"

    month = int(match.group(1))
    day = int(match.group(2))
    date_year = int(match.group(3))

    # Validate if the date actually exists
    try:
        datetime(date_year, month, day)
        # Log successful validation for debugging
        print(f"Valid date: '{date_str}'")
        return date_str
    except ValueError:
        print(
            f"Invalid date '{date_str}' (e.g., June 31st), replacing with 01/01/{year}"
        )
        return f"01/01/{year}"


# ===== Helper For Candidates Functions =====
def _download_candidate_xlsx(url: str, output_dir: Path, logger) -> str | None:
    """Download a candidate XLSX file and return the file path if successful."""
    try:
        # Use our custom library to download the XLSX file
        xlsx_filename = url.split("/")[-1]
        output_path = output_dir / xlsx_filename

        stream_download_file_to_path(
            request_url=url, file_save_path=output_path, headers=CANDIDATES_HEADERS
        )

        logger.info(f"Successfully downloaded: {xlsx_filename}")
        return str(output_path)

    except Exception as e:
        logger.warning(f"Failed to download xlsx file from {url}: {e}")
        return None


def _parse_candidate_table(
    page_content: str, year: int, logger, output_dir: Path, page_name: str
) -> str | None:
    """Parse candidate table from HTML content
    save to CSV, and return the file path
    """

    try:
        soup = BeautifulSoup(page_content, "html.parser")

        # Find the table with the specified class
        table = soup.find(
            "table", class_="table table-condensed table-bordered table-striped"
        )
        if not isinstance(table, Tag):
            logger.warning(f"No candidate table found for year {year}")
            return None

        rows = table.find_all("tr")[2:]

        candidates = []

        for row in rows:
            if not isinstance(row, Tag):
                continue

            cells = row.find_all("td")
            if len(cells) >= 6:
                if not all(isinstance(cell, Tag) for cell in cells[:6]):
                    continue

                name_address = cells[3] if cells[3] else ""
                phone = cells[4].get_text(strip=True) if cells[4] else ""
                filing_date = cells[5].get_text(strip=True) if cells[5] else ""

                if isinstance(name_address, Tag):
                    name_span = name_address.find("span", class_="local-note")
                    ballot_name = name_span.get_text(strip=True) if name_span else ""

                    address_lines = []
                    for br in name_address.find_all("br"):
                        if isinstance(br.previous_sibling, str):
                            text = br.previous_sibling.strip()
                            if text and not text.startswith(
                                ("Email:", "Url:", "Mailing Address:")
                            ):
                                address_lines.append(text)

                    address_line1 = address_lines[0] if len(address_lines) > 0 else ""
                    address_line2 = address_lines[1] if len(address_lines) > 1 else ""

                    email_link = name_address.find(
                        "a",
                        href=lambda x: isinstance(x, str) and x.startswith("mailto:"),
                    )
                    # Note: email_address is extracted but not currently used
                    # Keeping for potential future use
                    _ = email_link.get_text(strip=True) if email_link else ""
                else:
                    ballot_name = ""
                    address_line1 = ""
                    address_line2 = ""

                # Skip candidates without names
                if not ballot_name or ballot_name.strip() == "":
                    continue

                candidate = {
                    "name": ballot_name,
                    "address": f"{address_line1}; {address_line2}".strip("; "),
                    "party": "",
                    "date_filed": validate_and_fix_date(filing_date, year),
                    "phone": phone,
                }

                candidates.append(candidate)

        if not candidates:
            logger.warning(f"No candidate data could be parsed from {page_name}")
            return None

        # Save parsed data to CSV with same name as page
        csv_filename = f"{page_name}.csv"
        csv_path = output_dir / csv_filename

        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = candidates[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(candidates)

        logger.info(
            f"Successfully parsed {len(candidates)} "
            f"candidates and saved to {csv_filename}"
        )
        return str(csv_path)

    except Exception as e:
        logger.error(f"Error parsing candidate table for year {year}: {e}")
        return None


def _map_xlsx_columns_to_standard_headers(headers: list[str]) -> dict[str, int]:
    """
    Map XLSX column headers to our simplified DE_CANDIDATES_HEADERS.
    Returns a mapping of our standard header names to column indices.
    """
    mapping = {}

    # Clean headers for comparison
    clean_headers = [str(h).strip() for h in headers]

    # Use the exact working logic from index.py
    column_mapping_dict = {
        "name": ["BallotName"],
        "address": [
            "Residential_Address_Line_1",
            "Mailing_Address_Line_1",
            "Address_Line1",
            "Address_Line_1",
            "Address_Line2",
        ],
        "party": ["Party", "Fills_term_Balance", "Fills_Term_Balance"],
        "date_filed": ["Filing Date"],
        "phone": ["Phone_Number_1", "Phone_Number_2", "Phone_Number"],
    }

    # Find matching columns for each standard header
    for standard_header, possible_xlsx_headers in column_mapping_dict.items():
        mapping[standard_header] = -1  # Default to not found

        # Look for any of the possible XLSX headers that map to this standard header
        for xlsx_header in possible_xlsx_headers:
            for i, header in enumerate(clean_headers):
                if header == xlsx_header:  # Use exact string matching like index.py
                    mapping[standard_header] = i
                    break
            if mapping[standard_header] != -1:
                break

    # Special handling for address columns - we need both Line1 and Line2 if they exist
    address_line1_idx = -1
    address_line2_idx = -1

    for i, header in enumerate(clean_headers):
        if header == "Address_Line1" or header == "Address_Line_1":
            address_line1_idx = i
        elif header == "Address_Line2":
            address_line2_idx = i

    # Store both indices for address processing
    mapping["address_line1"] = address_line1_idx
    mapping["address_line2"] = address_line2_idx

    return mapping


def _parse_xlsx_candidate_data(file_path: Path, logger, year: int) -> list[dict]:
    """
    Parse XLSX candidate data and map to standardized format using openpyxl.
    Returns a list of candidate dictionaries with DE_CANDIDATES_HEADERS keys.
    """
    try:
        from openpyxl import load_workbook

        # Load the workbook
        workbook = load_workbook(filename=file_path, data_only=True)
        worksheet = workbook.active

        if worksheet is None:
            logger.error("No active worksheet found in the XLSX file")
            return []

        # Get column headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        logger.info(f"Found columns in XLSX: {headers}")

        # Map columns to our standard headers
        column_mapping = _map_xlsx_columns_to_standard_headers(headers)
        logger.info(f"Column mapping: {column_mapping}")

        # Log which columns were found and which were not
        found_columns = {k: v for k, v in column_mapping.items() if v >= 0}
        missing_columns = {k: v for k, v in column_mapping.items() if v == -1}
        logger.info(f"Found columns: {found_columns}")
        logger.info(f"Missing columns (will use empty strings): {missing_columns}")

        candidates = []

        # Check if worksheet has data
        if worksheet.max_row < 2:
            logger.warning("Worksheet has no data rows (only header)")
            return []

        # Process each row (skip header row) using the working logic from index.py
        for row_num in range(2, worksheet.max_row + 1):
            # Extract row data
            row_data = []
            for cell in worksheet[row_num]:
                row_data.append(cell.value)

            # Create candidate dictionary with default empty values
            candidate = dict.fromkeys(DE_CANDIDATES_HEADERS, "")

            # Special handling for address columns - join Line1 and Line2 if both exist
            address_line1_value = ""
            address_line2_value = ""

            if (
                column_mapping["address_line1"] >= 0
                and len(row_data) > column_mapping["address_line1"]
            ):
                address_line1_value = str(
                    row_data[column_mapping["address_line1"]] or ""
                ).strip()

            if (
                column_mapping["address_line2"] >= 0
                and len(row_data) > column_mapping["address_line2"]
            ):
                address_line2_value = str(
                    row_data[column_mapping["address_line2"]] or ""
                ).strip()

            # Join address lines if both have values
            if address_line1_value and address_line2_value:
                candidate["address"] = f"{address_line1_value}, {address_line2_value}"
            elif address_line1_value:
                candidate["address"] = address_line1_value
            elif address_line2_value:
                candidate["address"] = address_line2_value

            # Map data from XLSX columns to standard headers
            # (excluding address which we handled above)
            for idx, value in enumerate(row_data):
                if idx < len(headers):
                    # Find which standard header this column maps to
                    for standard_header, col_idx in column_mapping.items():
                        if col_idx == idx and standard_header not in [
                            "address_line1",
                            "address_line2",
                        ]:
                            if value is not None:
                                # Special handling for date_filed field -
                                # validate and fix dates
                                if standard_header == "date_filed":
                                    candidate[standard_header] = validate_and_fix_date(
                                        value, year
                                    )
                                else:
                                    # If multiple columns map to same field,
                                    # append with semicolon
                                    if (
                                        candidate[standard_header]
                                        and str(value).strip()
                                    ):
                                        candidate[standard_header] = (
                                            f"{candidate[standard_header]}; "
                                            f"{str(value).strip()}"
                                        )
                                    elif str(value).strip():
                                        candidate[standard_header] = str(value).strip()
                            break

            candidates.append(candidate)

        workbook.close()

        logger.info(f"Successfully parsed {len(candidates)} candidates from XLSX")
        if candidates:
            logger.info(f"Sample candidate data: {candidates[0]}")

        return candidates

    except Exception as e:
        logger.error(f"Error parsing XLSX file {file_path}: {e}")
        return []


# ===== Fetch Assets =====
@dg.asset(
    partitions_def=de_candidates_yearly_partition,
)
def de_fetch_candidates_data(context: dg.AssetExecutionContext):
    """
    Fetch candidates data from Delaware's elections website for a specific year.
    Scrapes the candidates page to find candidate list pages and downloads xlsx files.
    This asset is partitioned by year, fetching data for each year starting from 2008.
    """
    logger = context.log
    year = int(context.partition_key)

    # name = "candidates"  # Not used

    # Skip years before 2008
    if year < 2008:
        logger.info(
            f"Candidate data not available for year {year}\
            (before 2008), skipping"
        )
        return

    # Setup output directory
    output_dir = Path(DE_DATA_PATH_PREFIX) / "candidates" / str(year)
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Fetching candidates data for year {year}")

    # Fetch main candidates page using Playwright
    with sync_playwright() as p:
        browser = p.webkit.launch(headless=True)
        playwright_context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/138.0.0.0 Safari/537.36 Edg/138.0.0.0"
            )
        )

        main_page = playwright_context.new_page()

        # Fetch main candidates list page
        main_page.goto(
            "https://elections.delaware.gov/candidates/candidatelist/",
            wait_until="domcontentloaded",
            timeout=30000,
        )
        main_page.wait_for_timeout(2000)

        response_text = main_page.content()
        logger.info(f"Successfully fetched candidates page for year {year}")

        # Find candidate pages for this year
        candidate_pages = re.findall(get_candidate_page_pattern(year), response_text)
        logger.info(
            f"Found {len(candidate_pages)} candidate list pages for year {year}"
        )

        downloaded_files = []

        # Process each candidate page
        for page_url, page_name in candidate_pages:
            full_page_url = (
                f"{DE_CANDIDATES_BASE_URL}{page_url}"
                if page_url.startswith("/")
                else page_url
            )
            logger.info(f"Processing candidate page: {page_name}")

            # Navigate to candidate page
            main_page.goto(full_page_url, wait_until="domcontentloaded", timeout=30000)
            main_page.wait_for_timeout(1000)

            page_content = main_page.content()
            xlsx_links = re.findall(get_xlsx_file_pattern(year), page_content)

            if xlsx_links:
                xlsx_link, _ = xlsx_links[0]
                full_xlsx_url = (
                    f"{DE_CANDIDATES_BASE_URL}{xlsx_link}"
                    if xlsx_link.startswith("/")
                    else xlsx_link
                )

                logger.info(f"Downloading xlsx file from: {full_xlsx_url}")

                file_path = _download_candidate_xlsx(full_xlsx_url, output_dir, logger)

                if file_path:
                    # Process the XLSX file to standardize columns and save as CSV
                    logger.info(f"Processing downloaded XLSX file: {file_path}")
                    candidates_data = _parse_xlsx_candidate_data(
                        Path(file_path), logger, year
                    )

                    if candidates_data:
                        # Final data cleaning before saving to ensure CSV is clean
                        final_cleaned_candidates = []
                        for _i, candidate in enumerate(candidates_data):
                            final_candidate = {}
                            for header in DE_CANDIDATES_HEADERS:
                                value = candidate.get(header, "")

                                # Final cleanup for date fields
                                # ensure they're valid MM/DD/YYYY and fix invalid dates
                                if header == "date_filed":
                                    # Use the year from the partition context for
                                    # date validation
                                    final_candidate[header] = validate_and_fix_date(
                                        value, year
                                    )
                                else:
                                    final_candidate[header] = (
                                        str(value).strip() if value else ""
                                    )

                            final_cleaned_candidates.append(final_candidate)

                        # Save cleaned data as CSV with same name as XLSX file
                        xlsx_filename = Path(file_path).stem  # Get filename without ext
                        csv_filename = f"{xlsx_filename}.csv"
                        csv_path = output_dir / csv_filename

                        with open(
                            csv_path, "w", newline="", encoding="utf-8"
                        ) as csvfile:
                            fieldnames = DE_CANDIDATES_HEADERS
                            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                            writer.writeheader()
                            writer.writerows(final_cleaned_candidates)

                        # Replace the XLSX file path with the CSV path in
                        # downloaded_files
                        downloaded_files.append(str(csv_path))
                        logger.info(
                            f"Successfully standardized and saved "
                            f"{len(candidates_data)} candidates to {csv_filename}"
                        )

                        # Remove the original XLSX file since we now have the
                        # standardized CSV
                        # try:
                        #     Path(file_path).unlink()
                        #     logger.info(f"Removed original XLSX file: {file_path}")
                        # except Exception as e:
                        #     logger.warning(
                        #         f"Could not remove original XLSX file "
                        #         f"{file_path}: {e}"
                        #     )
                    else:
                        # If parsing failed, keep the original XLSX file
                        downloaded_files.append(file_path)
                        logger.warning(
                            f"Failed to parse XLSX file, keeping original: {file_path}"
                        )
            else:
                logger.info(
                    f"No xlsx file found for {page_name} - parsing table instead"
                )

                csv_path = _parse_candidate_table(
                    page_content, year, logger, output_dir, page_name
                )

                if csv_path:
                    downloaded_files.append(csv_path)
                    logger.info(
                        f"Successfully parsed and saved candidates to {csv_path}"
                    )
                else:
                    logger.warning(
                        f"No candidate data could be parsed from {page_name}"
                    )

        playwright_context.close()
        browser.close()

    logger.info(f"Downloaded {len(downloaded_files)} files for year {year}")


@dg.asset
def de_fetch_filed_reports_data():
    """
    Fetch filed reports data from Delaware's campaign finance system.
    Downloads CSV files for each committee type.
    """
    logger = dg.get_dagster_logger(name="de_filed_reports_fetch")

    try:
        name = "filed_reports"
        session = requests.Session()

        output_dir = Path(DE_DATA_PATH_PREFIX) / name
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Process each committee type
        for committee_type in COMMITTEE_TYPES:
            logger.info(f"Fetching data for committee type {committee_type}")

            # Prepare payload for POST request
            payload = {
                "hdnIsAdmin": "",
                "hdnViewCurrent": "",
                "hdnFilingPeriodName": "",
                "hdnTP": "",
                "txtCommitteeID": "",
                "MemberId": "",
                "txtCommitteeName": "",
                "hdnCommitteeName": "",
                "FilingPeriodName": "",
                "ReportName": "",
                "hdnReportName": "|",
                "dtStartDate": "",
                "dtEndDate": "",
                "dtCloseStartDate": "",
                "dtCloseEndDate": "",
                "ddlElectiontype": "",
                "ddljurisdiction": "",
                "CommitteeType": committee_type,
                "ddlOffice": "",
                "ddlCounty": "",
                "ddlOfficeSought": "",
                "btnSearch": "Search",
                "hdnddlOffice": "",
                "hdnddlCounty": "",
                "hdnddlOfficeSought": "",
                "hdnddljurisdiction": "",
            }

            # Make POST request to view reports
            response = session.post(
                DE_VIEW_REPORTS_URL,
                data=payload,
                headers=HEADERS,
                params={"theme": "vista"},
            )
            response.raise_for_status()
            logger.info(f"Fetched data for committee type {committee_type}")

            # Export to CSV
            csv_response = session.get(
                DE_EXPORT_CSV_URL,
                headers=HEADERS,
                params={"page": "1", "orderBy": "~", "filter": "~", "theme": "vista"},
            )
            csv_response.raise_for_status()

            # Save CSV file
            csv_filename = f"{output_dir}/type_{committee_type}_{name}.csv"
            with open(csv_filename, "wb") as f:
                f.write(csv_response.content)

            logger.info(
                f"Saved filed reports for committee\
                        type {committee_type} to {csv_filename}"
            )

    except Exception as e:
        logger.error(f"Error fetching Delaware filed reports: {e!s}")
        raise


@dg.asset
def de_fetch_committees_data():
    """
    Fetch committees data from Delaware's campaign finance system.
    Downloads CSV files for each office type.
    """
    logger = dg.get_dagster_logger(name="de_committees_fetch")

    try:
        name = "committees"
        session = requests.Session()

        output_dir = Path(DE_DATA_PATH_PREFIX) / name
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Process each office type
        for office_type in OFFICE_TYPES:
            logger.info(f"Fetching data for office type {office_type}")

            # Prepare payload for POST request
            payload = {
                "theme": "vista",
                "txtCommitteeID": "",
                "CommitteeType": "",
                "MemberId": "",
                "txtCommitteeName": "",
                "hdnAcronymId": "",
                "txtAcronym": "",
                "FormType": "",
                "CommitteeStatus": "",
                "hdnPersonID": "",
                "txtResOfficer": "",
                "ddlOffice": office_type,
                "ddlCounty": "",
                "ddlOfficeSought": "",
                "ddljurisdiction": "",
                "dtStartDate": "",
                "dtEndDate": "",
                "dtCloseStartDate": "",
                "dtCloseEndDate": "",
                "btnSearch": "Search",
                "hdnddlOffice": office_type,
                "hdnddlCounty": "",
                "hdnddlOfficeSought": "",
                "hdnddljurisdiction": "",
            }

            # Make POST request to search
            response = session.post(
                DE_SEARCH_URL, data=payload, headers=HEADERS, params={"theme": "vista"}
            )
            response.raise_for_status()
            logger.info(f"Fetched data for office type {office_type}")

            # Export to CSV
            csv_response = session.get(
                DE_EXPORT_COMMITTEES_URL,
                headers=HEADERS,
                params={"page": "1", "orderBy": "~", "filter": "~", "theme": "vista"},
            )
            csv_response.raise_for_status()

            # Save CSV file
            csv_filename = f"{output_dir}/{office_type}_committees.csv"
            with open(csv_filename, "wb") as f:
                f.write(csv_response.content)

            logger.info(
                f"Saved committees for office type\
                        {office_type} to {csv_filename}"
            )

    except Exception as e:
        logger.error(f"Error fetching Delaware committees: {e!s}")
        raise


@dg.asset
def de_fetch_expenditures_data():
    """
    Fetch expenditures data from Delaware's campaign finance system.
    Downloads CSV files for each committee type.
    """
    logger = dg.get_dagster_logger(name="de_expenditures_fetch")

    try:
        name = "expenditures"
        session = requests.Session()

        output_dir = Path(DE_DATA_PATH_PREFIX) / name
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Process each committee type
        for committee_type in COMMITTEE_TYPES:
            logger.info(f"Fetching data for committee type {committee_type}")

            # Prepare payload for POST request
            payload = {
                "theme": "vista",
                "hdnTP": "",
                "ddlPayeeType": "",
                "txtPayeeLastName": "",
                "txtPayeeFirstName": "",
                "txtStreet": "",
                "txtTown": "",
                "ddlState": "",
                "txtZipCode": "",
                "txtZipExt": "",
                "txtCommitteeID": "",
                "ddlExpenCategory": "",
                "ddlExpensePurpose": "",
                "MemberId": "",
                "txtRegistrant": "",
                "CommitteeType": committee_type,
                "dtStartDate": "",
                "dtEndDate": "",
                "txtAmountfrom": "",
                "txtAmountto": "",
                "hdnFixedAssets": "",
                "FilingYear": "",
                "FilingPeriodName": "",
                "Submit": "Search",
            }

            # Make POST request to other search
            response = session.post(
                DE_OTHER_SEARCH_URL,
                data=payload,
                headers=HEADERS,
                params={"theme": "vista"},
            )
            response.raise_for_status()
            logger.info(f"Fetched data for committee type {committee_type}")

            # Export to CSV
            csv_response = session.get(
                DE_EXPORT_EXPENSES_URL,
                headers=HEADERS,
                params={"page": "1", "orderBy": "~", "filter": "~"},
            )
            csv_response.raise_for_status()

            # Save CSV file
            csv_filename = f"{output_dir}/type_{committee_type}_{name}.csv"
            with open(csv_filename, "wb") as f:
                f.write(csv_response.content)

            logger.info(
                f"Saved expenditures for committee type\
                        {committee_type} to {csv_filename}"
            )

    except Exception as e:
        logger.error(f"Error fetching Delaware expenditures: {e!s}")
        raise


@dg.asset(
    partitions_def=de_contributions_yearly_partition,
)
def de_fetch_contributions_data(context: dg.AssetExecutionContext) -> None:
    """
    Fetch contributions data from Delaware's campaign finance system
    for a specific year.
    This asset is partitioned by year, fetching data for each year.
    """
    logger = context.log
    year = int(context.partition_key)

    try:
        # Initialize session
        session = requests.Session()

        # Create data directory
        name = "contributions"
        output_dir = Path(DE_DATA_PATH_PREFIX) / name / str(year)
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        logger.info(f"Fetching contributions data for year {year}")

        # Prepare payload for POST request
        payload = {
            "theme": "vista",
            "hdnTP": "",
            "txtContributorName": "",
            "txtFirstName": "",
            "txtStreet": "",
            "txtTown": "",
            "MemberId": "",
            "FilingYear": str(year),
            "FilingPeriodName": "",
            "ContributorType": "",
            "ContributionType": "",
            "ddlState": "",
            "txtZipCode": "",
            "txtZipExt": "",
            "dtStartDate": "",
            "dtEndDate": "",
            "txtAmountRangeFrom": "",
            "txtAmountRangeTo": "",
            "ddlOffice": "",
            "ddlCounty": "",
            "ddlOfficeSought": "",
            "ddljurisdiction": "",
            "txtReceivingRegistrant": "",
            "ddlEmployerOccupation": "",
            "hdnFixedAssets": "",
            "btnSearch": "Search",
            "hdnddlOffice": "",
            "hdnddlCounty": "",
            "hdnddlOfficeSought": "",
            "hdnddljurisdiction": "",
        }

        # Make POST request to view receipts
        response = session.post(
            DE_VIEW_RECEIPTS_URL,
            data=payload,
            headers=HEADERS,
            params={"theme": "vista"},
        )
        response.raise_for_status()
        logger.info(f"Fetched data for year {year}")

        # Export to CSV
        csv_response = session.get(
            DE_EXPORT_CONTRIBUTIONS_URL,
            headers=HEADERS,
            params={
                "page": "1",
                "orderBy": "~",
                "filter": "~",
                "Grid-size": "15",
                "theme": "vista",
            },
        )
        csv_response.raise_for_status()

        # Save CSV file
        csv_filename = f"{output_dir}/{year}_contributions.csv"
        with open(csv_filename, "wb") as f:
            f.write(csv_response.content)

        logger.info(f"Saved contributions for year {year} to {csv_filename}")

    except Exception as e:
        logger.error(f"Error fetching Delaware contributions for year {year}: {e!s}")
        raise


# ===== Insert Assets =====
@dg.asset(
    deps=[de_fetch_candidates_data],
    partitions_def=de_candidates_yearly_partition,
)
def de_insert_candidates_to_landing_table(
    context: dg.AssetExecutionContext, pg: dg.ResourceParam[PostgresResource]
):
    """
    Insert candidates data into the landing table for a specific year.
    Deletes existing data for the year before inserting new data.
    """
    logger = context.log
    year = int(context.partition_key)

    # Get all candidate files for the specific year
    data_dir = Path(DE_DATA_PATH_PREFIX) / "candidates" / str(year)
    candidate_files = list(data_dir.glob("*.csv")) + list(data_dir.glob("*.xlsx"))

    if not candidate_files:
        logger.info(f"No candidate files found for year {year}")
        return dg.MaterializeResult(
            metadata={
                "dagster/table_name": "de_candidates_landing",
                "dagster/row_count": 0,
                "dagster/year": year,
                "dagster/status": "no_files_found",
            }
        )

    table_name = "de_candidates_landing"
    table_name_identifier = sql.Identifier(table_name)

    # Create a truncate query that only deletes data for the specific year
    # Use string functions instead of TO_DATE to avoid crashes on invalid dates
    truncate_query = sql.SQL(
        "DELETE FROM {table_name} WHERE SPLIT_PART(date_filed, '/', 3) = {year}"
    ).format(table_name=table_name_identifier, year=sql.Literal(str(year)))

    return de_insert_raw_file_to_landing_table(
        postgres_pool=pg.pool,
        table_name=table_name,
        table_columns_name=DE_CANDIDATES_HEADERS,
        data_validation_callback=lambda row: len(row) == len(DE_CANDIDATES_HEADERS),
        category="candidates",
        data_files=[str(f) for f in candidate_files],
        truncate_query=truncate_query,
    )


@dg.asset(deps=[de_fetch_filed_reports_data])
def de_insert_filed_reports_to_landing_table(pg: dg.ResourceParam[PostgresResource]):
    """
    Insert filed reports data into the landing table.
    """
    return de_insert_raw_file_to_landing_table(
        postgres_pool=pg.pool,
        table_name="de_filed_reports_landing",
        table_columns_name=DE_FILED_REPORTS_HEADERS,
        data_validation_callback=lambda row: len(row) == len(DE_FILED_REPORTS_HEADERS),
        category="filed_reports",
    )


@dg.asset(deps=[de_fetch_committees_data])
def de_insert_committees_to_landing_table(pg: dg.ResourceParam[PostgresResource]):
    """
    Insert committees data into the landing table.
    """
    return de_insert_raw_file_to_landing_table(
        postgres_pool=pg.pool,
        table_name="de_committees_landing",
        table_columns_name=DE_COMMITTEES_HEADERS,
        data_validation_callback=lambda row: len(row) == len(DE_COMMITTEES_HEADERS),
        category="committees",
    )


@dg.asset(deps=[de_fetch_expenditures_data])
def de_insert_expenditures_to_landing_table(pg: dg.ResourceParam[PostgresResource]):
    """
    Insert expenditures data into the landing table.
    """
    return de_insert_raw_file_to_landing_table(
        postgres_pool=pg.pool,
        table_name="de_expenditures_landing",
        table_columns_name=DE_EXPENDITURES_HEADERS,
        data_validation_callback=lambda row: len(row) == len(DE_EXPENDITURES_HEADERS),
        category="expenditures",
    )


@dg.asset(
    deps=[de_fetch_contributions_data],
    partitions_def=de_contributions_yearly_partition,
)
def de_insert_contributions_to_landing_table(
    context: dg.AssetExecutionContext, pg: dg.ResourceParam[PostgresResource]
):
    """
    Insert contributions data into the landing table for a specific year.
    """
    year = int(context.partition_key)

    # Get all contribution CSV files for the specific year
    data_dir = Path(DE_DATA_PATH_PREFIX) / "contributions" / str(year)
    contribution_files = list(data_dir.glob(f"{year}_contributions.csv"))

    if not contribution_files:
        return dg.MaterializeResult(metadata={"dagster/row_count": 0})

    table_name = "de_contributions_landing"
    table_name_identifier = sql.Identifier(table_name)
    # Create a truncate query that only deletes data for the specific year
    truncate_query = sql.SQL(
        "DELETE FROM {table_name} WHERE "
        "EXTRACT(YEAR FROM TO_DATE(contribution_date, 'MM/DD/YYYY')) = {year}"
    ).format(table_name=table_name_identifier, year=sql.Literal(year))

    return de_insert_raw_file_to_landing_table(
        postgres_pool=pg.pool,
        table_name=table_name,
        table_columns_name=DE_CONTRIBUTIONS_HEADERS,
        data_validation_callback=lambda row: len(row) == len(DE_CONTRIBUTIONS_HEADERS),
        category="contributions",
        data_files=[str(f) for f in contribution_files],
        truncate_query=truncate_query,
    )
